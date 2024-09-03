from datetime import datetime

from openpyxl.workbook import Workbook

from main import _add_headers, HEADERS, _add_content
from parser import Row

sample_rows = [
    Row(
        date=datetime(2023, 8, 1),
        description="Apples",
        installment="No",
        amount=1.50,
    ),
    Row(
        date=datetime(2023, 8, 2),
        description="Bananas",
        installment="1/2",
        amount=0.80,
    ),
]


def test_add_headers():
    wb = Workbook()
    ws = wb.active

    _add_headers(ws)

    assert [cell.value for cell in ws[1]] == HEADERS


def test_add_content():
    wb = Workbook()
    ws = wb.active

    _add_headers(ws)
    _add_content(ws, sample_rows)

    # Check the content values
    for row, sample_row in zip(ws.iter_rows(min_row=2, values_only=True), sample_rows):
        assert row == (
            sample_row.date,
            sample_row.description,
            sample_row.amount,
            "",
            "",
            sample_row.installment,
        )
